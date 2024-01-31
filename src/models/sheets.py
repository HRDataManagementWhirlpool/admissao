import os
from openpyxl import load_workbook

class SheetsModel:
    def __init__(self, folder_path, file_keyword):
        self.folder_path = folder_path
        self.file_keyword = file_keyword

    def load_sheet(self):
            for file in os.listdir(self.folder_path):
                if self.file_keyword in file:
                    file_path = os.path.join(self.folder_path, file)
                    workbook = load_workbook(filename=file_path, data_only=True).active
                    print(f'{self.file_keyword} loaded')
                    return workbook
            return None
    
    def clone_sheet(self, worksheet):
        for file in os.listdir(self.folder_path):
            if self.file_keyword in file:
                file_path = os.path.join(self.folder_path, file)
                workbook = load_workbook(filename=file_path)
                worksheet = workbook[worksheet]
                print(f'{self.file_keyword} loaded')
                return worksheet, workbook