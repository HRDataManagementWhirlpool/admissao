import os
import datetime
from openpyxl import load_workbook, Workbook
from docusign import DocusignController

class SheetsController:
    def start_process(checkList, conferencia, contas, dependentes, eSocial, workForce, folder_path, check):
        resConferidos = [] # Armazena o RE para evitar duplicidade
        for re in checkList['B'][2:]: # Para linha na Coluna B a partir da terceira linha:
            nomesConferidos = [] # Armazena os nomes dos dependentes do RE que está sendo conferido para evitar duplicidade
            if re.value is not None: # Caso o valor do RE não seja vazio
                for celula in conferencia['A'][1:]: # Verifica se o valor do RE a ser conferido é igual ao RE dentro da tabela
                    if celula.value == re.value: # Caso seja igual
                        if re.value not in resConferidos: # E não esteja duplicado
                            resConferidos.append(re.value) # Insere o RE na lista
                            array = SheetsController.get_data(celula.row, conferencia, contas, dependentes, eSocial, workForce, re.value, nomesConferidos) # Armazena os valores em uma lista
                            SheetsController.fill_sheet(check, re.row, array) # Preenche as informações na planilha
        SheetsController.savefile(folder_path, check) # Salva o arquivo e finaliza o processo

    def format_date(date_string):
        return datetime.date(int(date_string[:4]), int(date_string[5:7]), int(date_string[8:10]))

    def workbook_data(workbook, init, reValue, col):
        for cell in workbook[init][1:]:
            row = cell.row
            if reValue in str(cell.value):
                data = str(workbook[f'{col}{row}'].value)
                return data

    def get_data(linha, conferencia, contas, dependentes, eSocial, workForce, reValue, nomesConferidos):
        dtInicio = SheetsController.format_date(str(conferencia[f'C{linha}'].value))
        vinculo = conferencia[f'AF{linha}'].value
        codCategoria = conferencia[f'AL{linha}'].value
        codVinculo = conferencia[f'AM{linha}'].value
        codExposicao = conferencia[f'AN{linha}'].value
        agencia = SheetsController.workbook_data(contas, 'A', reValue, 'D')[-4:]
        conta = SheetsController.workbook_data(contas, 'A', reValue, 'E')
        dependente = 0
        for cell in dependentes['A'][1:]:
            row = cell.row
            if cell.value == reValue:
                nome = dependentes[f'L{row}'].value
                if nome not in nomesConferidos:
                    dependente+=1
                    nomesConferidos.append(nome)
        saude = conferencia[f'P{linha}'].value
        saudeDt = SheetsController.format_date(str(conferencia[f'S{linha}'].value))
        vida = conferencia[f'V{linha}'].value
        vidaDt = SheetsController.format_date(str(conferencia[f'X{linha}'].value))
        fgts = conferencia[f'AG{linha}'].value
        pFgts = conferencia[f'AH{linha}'].value
        infotipo = '-'
        turno = conferencia[f'K{linha}'].value
        cargaHoraria = str(conferencia[f'J{linha}'].value)
        cargaHoraria = f'{cargaHoraria} hs'
        salario = conferencia[f'AE{linha}'].value
        sindicato = conferencia[f'AB{linha}'].value
        gremio='-'
        if conferencia[f'A{linha}'].value == conferencia[f'A{linha+1}'].value:
            gremio = conferencia[f'AB{linha+1}'].value
        periculosidade = conferencia[f'AK{linha}'].value
        
        es = SheetsController.workbook_data(eSocial, 'J', reValue, 'G')
        wf = SheetsController.workbook_data(workForce, 'BD', reValue, 'CB')
        cargo = SheetsController.workbook_data(workForce, 'BD', reValue, 'BJ')
        
        varArray = [
                            dtInicio,
                            cargo,
                            vinculo,
                            codCategoria,
                            codVinculo,
                            codExposicao,
                            agencia,
                            conta,
                            dependente,
                            saude,
                            saudeDt,
                            vida,
                            vidaDt,
                            fgts,
                            pFgts,
                            infotipo,
                            turno,
                            cargaHoraria,
                            salario,
                            gremio,
                            sindicato,
                            periculosidade,
                            es,
                            wf
                        ]
        return varArray

    def fill_sheet(worksheet, row, array):
        for col in worksheet['Conferência'][f'AD{row}':f'BA{row}']:
            for i, cell in enumerate(col):
                cell.value = array[i]

    def savefile(folder_path, workbook):
        now = datetime.datetime.now().strftime("%d-%m")
        filename = f'Update-{now}.xlsx'
        savefile = os.path.join(folder_path, filename)
        workbook.save(savefile)

    def fill_mensalistas_data(workbook):
        mensalistas = [] # Cria o array das listas com os dados de cada mensalista
        wb = load_workbook(filename=workbook, data_only= True) # Carrega a planilha com as informações das admissões
        ws = wb.active # Escolhe a aba ativa da planilha
        conf=[] # Cria uma lista que armazenará os REs dos mensalistas para que não haja duplicidade nos dados
        for type in ws['AR'][1:]:
            if type.value == "MENSAL.":
                nome = ws[f'B{type.row}'].value
                re = ws[f'A{type.row}'].value
                email = ws[f'Q{type.row}'].value
                if re not in conf: # Utilizando o array para evitar duplicidade
                    mensalistas.append({'nome': nome, 're': re, 'email': email}) # Armazena os dados na lista
                    conf.append(re)
        return mensalistas

    def save_docusign_data(array:list):
        try:
            wb = load_workbook('data/docusign_data.xlsx')
        except:
            wb = Workbook()
        finally:
            ws = wb.active
            ws['A1']= "RE"
            ws['B1']= "NOME COMPLETO"
            ws['C1']= "EMAIL"
            ws['D1']= "ENVIADO"
            ws['E1']= "DATA DE ENVIO"
            ws['F1']= "COLETADO"
            ws['G1']= "DATA DE COLETA"
        
        row = ws.max_row+1
        array = array + ["Ok", datetime.date.today(), "NOk", "-"]
        for i, data in enumerate(array):
            ws.cell(column=1+i, row=row, value=data)
        wb.save('data/docusign_data.xlsx')
        
    def get_signed_contracts(login, senha):
        try:
            wb = load_workbook('data/docusign_data.xlsx')
        except Exception as error:
            return error
        ws = wb.active
        processo = DocusignController(login, senha)
        processo.open_website()
        processo.login()
        for coletado in ws['F'][1:]:
            if coletado.value == 'NOk':
                row = coletado.row
                nome = ws[f'B{row}'].value
                re = ws[f'A{row}'].value
                contrato = f'CONTRATO DE TRABALHO - {nome} - {re}'
                coleta = processo.get_signed_document(contrato)
                if coleta:
                    ws[f'F{row}'].value = "Ok"
                    ws[f'G{row}'].value = datetime.date.today()
        wb.save('data/docusign_data.xlsx')